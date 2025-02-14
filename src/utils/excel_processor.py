from openpyxl import load_workbook, Workbook
from models.out_format import OutFormat
from openpyxl.styles import Font
import shutil
import os
from pathlib import Path
from openpyxl import load_workbook, Workbook

class ExcelProcessor: 
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None 

    
    def read_excel(self):
        try:
            # Cargar el libro de trabajo
            self.workbook = load_workbook(self.file_path)
            # Seleccionar la hoja activa
            self.sheet = self.workbook.active
        except Exception as e:
            print(f"Error al cargar el archivo Excel: {e}")

    def write_excel(self, out_formats, output_path):
        try:
            # Verificar si el archivo existe
            if os.path.exists(output_path):
                workbook = load_workbook(output_path)  # Cargar archivo existente
            else:
                workbook = Workbook()  # Crear nuevo archivo
                sheet = workbook.active  # Crear hoja
                self._add_headers(sheet)  # Agregar cabeceras
                
            sheet = workbook.active
            sheet.title = "FORMATO_CODBARRA"
            # Agregar los datos procesados desde la fila 2
            for row_num, format_obj in enumerate(out_formats, start=2):
                sheet.cell(row=row_num, column=1, value=format_obj.code)
                sheet.cell(row=row_num, column=2, value=format_obj.amount)  
                sheet.cell(row=row_num, column=3, value=format_obj.id_local) 

            # Guardar el archivo
            workbook.save(output_path)
            print(f"Archivo generado correctamente en: {output_path}")

        except Exception as e:
            print(f"Error al escribir el archivo Excel: {e}")

    def _add_headers(self, sheet):
        headers = ["CODIGO_BARRA", "CANTIDAD", "ID_SUCURSAL"]
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)  # Cabecera en negrita

    def get_sheet(self):
        return self.sheet
